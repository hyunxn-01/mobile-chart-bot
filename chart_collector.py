"""
한국 App Store 모바일 게임 매출 차트 일간 수집·캘린더 기반 이동평균 분석·메일 자동화.
GitHub Actions / cron-job.org 트리거로 매일 한국 시간 오전 7시 37분 실행.

활성 시간축 (그날 추가되는 분석):
  - 매일: 1일선 (어제 vs 오늘)
  - 월요일: + 1주선 (전전주 vs 전주, 월~일 단위)
  - 매월 1일: + 1달선 (전전월 vs 전월, 캘린더 월)
  - 분기 시작일(1/1, 4/1, 7/1, 10/1): + 분기선 (전전분기 vs 전분기)
  - 1월 1일: + 1년선 (재작년 vs 작년)

Claude API 호출은 529 과부하 등 일시 오류 시 자동 재시도. 최종 실패해도 메일은 발송.
AI 분석 모델은 파일 상단 CLAUDE_MODEL 상수로 관리.
"""

import hashlib
import json
import os
import smtplib
import time
from datetime import datetime, timedelta
from pathlib import Path
from email.mime.multipart import MIMEMultipart
from email.mime.text import MIMEText
from email.mime.base import MIMEBase
from email import encoders

import pandas as pd
import requests
from openpyxl import Workbook
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.styles import Font, Alignment, PatternFill
from anthropic import Anthropic
from trusted_sources import TRUSTED_DOMAINS, is_trusted_source   # #11: 신뢰 출처 화이트리스트

ANTHROPIC_API_KEY = os.environ.get('ANTHROPIC_API_KEY')
GMAIL_USER = os.environ.get('GMAIL_USER')
GMAIL_APP_PASSWORD = os.environ.get('GMAIL_APP_PASSWORD')
RECIPIENT_EMAIL = os.environ.get('RECIPIENT_EMAIL')

# === AI 분석 모델 (변경 시 이 한 줄만 수정) ===
CLAUDE_MODEL = 'claude-opus-4-8'
# 게임명 한글 변환용 모델(저비용·캐시). 음차/표기 변환은 호출이 작고 결과는 캐시돼 비용 미미.
TRANSLATE_MODEL = 'claude-sonnet-4-6'
# 적응형 사고 강도: high(기본)=거의 항상 깊게 사고하되 max보다 사고토큰 절감(비용↓, 품질 거의 동일). 값: low/medium/high/max.
THINKING_EFFORT = 'high'
# 출력 토큰 상한(사고+응답 합산). 사고가 길어도 응답이 잘리지 않게 넉넉히.
MAX_OUTPUT_TOKENS = 32000

# === 대시보드 URL (GitHub Pages 주소) ===
# 형식: https://<github-사용자명>.github.io/<레포명>/
DASHBOARD_URL = 'https://hyunxn-01.github.io/mobile-chart-bot/'

DATA_DIR = Path('data')
DATA_DIR.mkdir(exist_ok=True)
HISTORY_DIR = DATA_DIR / 'history'
HISTORY_DIR.mkdir(exist_ok=True)
HISTORY_FREE_DIR = DATA_DIR / 'history_free'   # Top Free(인기) 차트 누적 — 대시보드 보조
HISTORY_FREE_DIR.mkdir(exist_ok=True)

# === 다국가 수집(iOS App Store 스토어프런트) ===
# 1차: T1 코어 10개국. 안정 후 EXTRA를 COUNTRIES에 합쳐 32개국으로 확장.
# 현재 수집(주요 10개국). 나머지 22개국은 사이트 완성 후 맨 마지막에 확장(저장 메타-분리 #98 후).
COUNTRIES = ['kr', 'us', 'jp', 'cn', 'tw', 'gb', 'de', 'fr', 'ca', 'au']
COUNTRIES_EXTRA = ['it', 'es', 'nl', 'ru', 'se', 'sa', 'ae', 'eg', 'tr',
                   'br', 'mx', 'ar', 'co', 'id', 'th', 'vn', 'ph', 'my', 'sg', 'in', 'pk', 'bd']
PRIMARY_COUNTRY = 'kr'                 # 기존 일일 메일·AI 브리핑 기준 국가(현행 유지)
CHARTS_DIR = DATA_DIR / 'charts'       # data/charts/{country}/{grossing|free}/{date}.json
CHARTS_DIR.mkdir(exist_ok=True)


# ============================================================
# 1. 데이터 수집·저장·로드
# ============================================================

def fetch_apple_chart_kr_games(limit=100):
    """Apple App Store 한국 게임 차트 수집. Top Grossing → Top Free fallback."""
    charts_to_try = [
        ('Top Grossing', f'https://itunes.apple.com/kr/rss/topgrossingapplications/limit={limit}/genre=6014/json'),
        ('Top Free', f'https://itunes.apple.com/kr/rss/topfreeapplications/limit={limit}/genre=6014/json'),
    ]
    for chart_name, url in charts_to_try:
        try:
            r = requests.get(url, timeout=15, headers={'User-Agent': 'Mozilla/5.0'})
            r.raise_for_status()
            entries = r.json().get('feed', {}).get('entry', [])
            if not entries:
                print(f"[WARN] {chart_name}: 빈 데이터. fallback.")
                continue
            apps = [
                {
                    'rank': i + 1,
                    'app_id': e.get('id', {}).get('attributes', {}).get('im:bundleId', ''),
                    'track_id': e.get('id', {}).get('attributes', {}).get('im:id', ''),
                    'title': e.get('im:name', {}).get('label', ''),
                    'developer': e.get('im:artist', {}).get('label', ''),
                    'category': e.get('category', {}).get('attributes', {}).get('label', ''),
                    'platform': 'App Store',
                    'chart': chart_name,
                }
                for i, e in enumerate(entries)
            ]
            print(f"[OK] {chart_name}: {len(apps)}개 수집")
            return apps, chart_name
        except Exception as e:
            print(f"[ERROR] {chart_name} 수집 실패: {e}")
    return [], None


def fetch_top_free_kr_games(limit=100):
    """Top Free(인기/다운로드) 차트 — 대시보드 보조용(획득 지표). 실패해도 메인 흐름엔 영향 없음."""
    url = f'https://itunes.apple.com/kr/rss/topfreeapplications/limit={limit}/genre=6014/json'
    try:
        r = requests.get(url, timeout=15, headers={'User-Agent': 'Mozilla/5.0'})
        r.raise_for_status()
        entries = r.json().get('feed', {}).get('entry', [])
        apps = [
            {
                'rank': i + 1,
                'app_id': e.get('id', {}).get('attributes', {}).get('im:bundleId', ''),
                'track_id': e.get('id', {}).get('attributes', {}).get('im:id', ''),
                'title': e.get('im:name', {}).get('label', ''),
                'developer': e.get('im:artist', {}).get('label', ''),
                'category': e.get('category', {}).get('attributes', {}).get('label', ''),
                'platform': 'App Store',
                'chart': 'Top Free',
            }
            for i, e in enumerate(entries)
        ]
        print(f"[OK] Top Free: {len(apps)}개 수집")
        return apps
    except Exception as e:
        print(f"[ERROR] Top Free 수집 실패: {e}")
        return []


# 게임 하위 장르(한국 스토어 표기). lookup genres 배열에서 이 중 첫 매칭을 대표 장르로.
GAME_SUBGENRES = ['롤플레잉', '전략', '시뮬레이션', '액션', '어드벤처', '퍼즐', '캐주얼',
                  '보드', '카드', '카지노', '스포츠', '레이싱', '아케이드', '가족', '단어', '트리비아', '음악']


def _pick_genre(genres):
    """genres(예: ['게임','롤플레잉','어드벤처'])에서 대표 장르 하나. 게임 하위 장르 우선, 없으면 '게임' 외 첫 값."""
    if not genres:
        return '미상'
    for g in genres:
        if g in GAME_SUBGENRES:
            return g
    for g in genres:
        if g != '게임':
            return g
    return '게임'


def fetch_genres(track_ids, country='kr'):
    """trackId 목록 → {trackId(str): 메타}. iTunes lookup(해당 country)의 genres·genreIds 사용. 실패 시 가능한 만큼만."""
    result = {}
    ids = [str(t) for t in track_ids if t]
    if not ids:
        return result
    for i in range(0, len(ids), 180):
        chunk = ids[i:i + 180]
        try:
            url = f"https://itunes.apple.com/lookup?id={','.join(chunk)}&country={country}"
            r = requests.get(url, timeout=20, headers={'User-Agent': 'Mozilla/5.0'})
            r.raise_for_status()
            for it in r.json().get('results', []):
                tid = str(it.get('trackId', ''))
                if tid:
                    result[tid] = {
                        'track_id': tid,
                        'version': it.get('version', ''),
                        'genre': _pick_genre(it.get('genres', [])),
                        'genres_all': ', '.join(it.get('genres', []) or []),
                        'genre_ids': ','.join(str(g) for g in (it.get('genreIds') or [])),
                        'primary_genre': it.get('primaryGenreName', ''),
                        'release': (it.get('releaseDate') or '')[:10],
                        'rating': it.get('averageUserRating'),
                        'icon': it.get('artworkUrl100', ''),
                        'updated': (it.get('currentVersionReleaseDate') or '')[:10],
                        'ratings': it.get('userRatingCount'),
                        'notes': (it.get('releaseNotes') or '')[:500],
                        'artist_id': it.get('artistId'),
                        'cv_rating': it.get('averageUserRatingForCurrentVersion'),
                        'cv_ratings': it.get('userRatingCountForCurrentVersion'),
                    }
        except Exception as e:
            print(f"[WARN] 장르 lookup 실패(chunk {i}): {e}")
    return result


def attach_genres(apps, country='kr'):
    """apps 각 게임에 'genre'·메타 추가. 실패해도 메인 흐름 무영향('미상')."""
    try:
        gmap = fetch_genres([a.get('track_id') for a in apps], country)
        for a in apps:
            m = gmap.get(str(a.get('track_id')), {})
            a['genre'] = m.get('genre', '미상')
            a['version'] = m.get('version', '')
            if m.get('track_id'):
                a['track_id'] = m.get('track_id')
            a['genres_all'] = m.get('genres_all', '')
            a['genre_ids'] = m.get('genre_ids', '')
            a['primary_genre'] = m.get('primary_genre', '')
            a['release'] = m.get('release', '')
            a['rating'] = m.get('rating')
            a['icon'] = m.get('icon', '')
            a['updated'] = m.get('updated', '')
            a['ratings'] = m.get('ratings')
            a['notes'] = m.get('notes', '')
            a['artist_id'] = m.get('artist_id')
            a['cv_rating'] = m.get('cv_rating')
            a['cv_ratings'] = m.get('cv_ratings')
        kinds = len({a.get('genre') for a in apps})
        print(f"[OK] 장르 부착: {len(apps)}개 게임 → {kinds}종 장르")
    except Exception as e:
        print(f"[WARN] 장르 부착 실패: {e}")
        for a in apps:
            a.setdefault('genre', '미상')
    return apps


def save_current_data(data):
    today = datetime.now().strftime('%Y-%m-%d')
    f = HISTORY_DIR / f'{today}.json'
    f.write_text(json.dumps(data, ensure_ascii=False, indent=2), encoding='utf-8')
    print(f"[OK] 데이터 저장: {f}")


def save_free_data(data):
    today = datetime.now().strftime('%Y-%m-%d')
    f = HISTORY_FREE_DIR / f'{today}.json'
    f.write_text(json.dumps(data, ensure_ascii=False, indent=2), encoding='utf-8')
    print(f"[OK] Top Free 저장: {f}")


def load_data_by_date(date_str):
    f = HISTORY_DIR / f'{date_str}.json'
    if f.exists():
        return json.loads(f.read_text(encoding='utf-8'))
    return None


def find_most_recent_past_data():
    today = datetime.now().strftime('%Y-%m-%d')
    files = sorted(HISTORY_DIR.glob('*.json'))
    past = [f for f in files if f.stem < today]
    if not past:
        return None, None
    f = past[-1]
    return json.loads(f.read_text(encoding='utf-8')), f.stem


def fetch_apple_chart(country, slug, limit=100):
    """임의 국가·차트(slug: topgrossingapplications/topfreeapplications) iOS 게임 차트 수집."""
    chart_name = 'Top Grossing' if 'grossing' in slug else 'Top Free'
    url = f'https://itunes.apple.com/{country}/rss/{slug}/limit={limit}/genre=6014/json'
    try:
        r = requests.get(url, timeout=15, headers={'User-Agent': 'Mozilla/5.0'})
        r.raise_for_status()
        entries = r.json().get('feed', {}).get('entry', [])
        return [
            {
                'rank': i + 1,
                'app_id': e.get('id', {}).get('attributes', {}).get('im:bundleId', ''),
                'track_id': e.get('id', {}).get('attributes', {}).get('im:id', ''),
                'title': e.get('im:name', {}).get('label', ''),
                'developer': e.get('im:artist', {}).get('label', ''),
                'platform': 'App Store',
                'chart': chart_name,
            }
            for i, e in enumerate(entries)
        ]
    except Exception as e:
        print(f"[ERROR] {country}/{slug} 수집 실패: {e}")
        return []


def fetch_titles(track_ids, country='kr'):
    """trackId 목록 → {trackId(str): trackName}. 해당 country 스토어의 표기명(스토어에 없으면 결과에서 빠짐)."""
    out = {}
    ids = [str(t) for t in track_ids if t]
    if not ids:
        return out
    for i in range(0, len(ids), 180):
        chunk = ids[i:i + 180]
        try:
            url = f"https://itunes.apple.com/lookup?id={','.join(chunk)}&country={country}"
            r = requests.get(url, timeout=20, headers={'User-Agent': 'Mozilla/5.0'})
            r.raise_for_status()
            for it in r.json().get('results', []):
                tid = str(it.get('trackId', ''))
                nm = it.get('trackName')
                if tid and nm:
                    out[tid] = nm
        except Exception as e:
            print(f"[WARN] 타이틀 lookup 실패(chunk {i}, {country}): {e}")
        time.sleep(0.3)
    return out


ALIASES_PATH = DATA_DIR / 'title_aliases.json'


def _needs_kr(s):
    """제목에 중국어 한자/일본어 가나가 있으면(=한국어·영문 아님) 변환 대상."""
    for ch in str(s or ''):
        o = ord(ch)
        if 0x3040 <= o <= 0x30FF or 0x3400 <= o <= 0x9FFF:  # 가나 + CJK 한자
            return True
    return False


def load_title_aliases():
    try:
        return json.loads(ALIASES_PATH.read_text(encoding='utf-8')) if ALIASES_PATH.exists() else {}
    except Exception:
        return {}


def save_title_aliases(aliases):
    try:
        ALIASES_PATH.write_text(json.dumps(aliases, ensure_ascii=False, sort_keys=True, indent=2), encoding='utf-8')
    except Exception as e:
        print(f"[WARN] 게임명 별칭 저장 실패: {e}")


def translate_titles_kr(titles):
    """현지어(중/일) 게임명 리스트 → {원문: 한국어명}. 50개씩 배치(응답 잘림 방지), 저비용 모델. 실패 배치만 건너뜀."""
    out = {}
    titles = list(titles)
    if not titles:
        return out
    client = Anthropic(api_key=ANTHROPIC_API_KEY)
    BATCH = 50
    for i in range(0, len(titles), BATCH):
        chunk = titles[i:i + BATCH]
        listing = '\n'.join(f"- {t}" for t in chunk)
        prompt = ("다음은 App Store 게임 제목(중국어 또는 일본어)들이다. 각 제목을 한국 게이머가 부르는 한국어 표기로 바꿔라. "
                  "한국 정식 서비스명이 있으면 그 이름을, 없으면 한글 음차로. 뜻 번역이 아니라 '게임 이름'만. "
                  "잘 모르면 한글 음차라도 반드시 한국어로. 출력은 JSON 객체 {\"원문\":\"한국어\"} 하나만, 다른 말 없이.\n\n" + listing)
        try:
            msg = client.messages.create(model=TRANSLATE_MODEL, max_tokens=4000,
                                         messages=[{'role': 'user', 'content': prompt}])
            txt = next((b.text for b in msg.content if getattr(b, 'type', None) == 'text'), '') or ''
            s, e = txt.find('{'), txt.rfind('}')
            if s >= 0 and e > s:
                out.update({str(k): str(v) for k, v in json.loads(txt[s:e + 1]).items() if v})
        except Exception as ex:
            print(f"[WARN] 게임명 번역 배치 실패(i={i}): {ex}")
        time.sleep(0.3)
    return out


def localize_titles(collected):
    """게임명 한국어화: KR 스토어명 → 없으면 US(영문) → 그래도 현지어(중/일)면 캐시된 별칭 → 없으면 저비용 AI 음차.
    각 app에 title_kr 부여."""
    all_ids = sorted({str(a.get('track_id')) for cc in collected for kind in collected[cc]
                      for a in collected[cc][kind] if a.get('track_id')})
    if not all_ids:
        return
    kr_titles = fetch_titles(all_ids, 'kr')
    missing = [i for i in all_ids if i not in kr_titles]
    us_titles = fetch_titles(missing, 'us') if missing else {}
    for cc in collected:
        for kind in collected[cc]:
            for a in collected[cc][kind]:
                tid = str(a.get('track_id'))
                a['title_kr'] = kr_titles.get(tid) or us_titles.get(tid) or a.get('title', '')
    # 여전히 중/일 현지어인 이름 → 캐시 우선, 새 이름만 저비용 모델로 음차(배치)
    try:
        aliases = load_title_aliases()
        pending = sorted({a['title_kr'] for cc in collected for kind in collected[cc]
                          for a in collected[cc][kind]
                          if _needs_kr(a.get('title_kr')) and a.get('title_kr') not in aliases})
        new_map = translate_titles_kr(pending) if pending else {}
        if new_map:
            aliases.update({k: v for k, v in new_map.items() if v and v != k})
            save_title_aliases(aliases)
        if aliases:
            for cc in collected:
                for kind in collected[cc]:
                    for a in collected[cc][kind]:
                        tk = a.get('title_kr', '')
                        if _needs_kr(tk) and tk in aliases:
                            a['title_kr'] = aliases[tk]
    except Exception as ex:
        print(f"[WARN] 게임명 별칭 단계 실패: {ex}")
    print(f"[OK] 게임명 현지화: KR {len(kr_titles)} + US {len(us_titles)} / 전체 {len(all_ids)}")


def apply_aliases_to_briefs(aliases=None):
    """기존 브리핑(major_*/region_*/global) 텍스트의 현지어 게임명을 한국어 별칭으로 치환.
    브리핑 재생성(Opus) 없이 문자열 치환만 — 무비용. 매 실행 idempotent."""
    aliases = aliases if aliases is not None else load_title_aliases()
    if not aliases:
        return
    out_dir = Path('docs') / 'markets'
    if not out_dir.exists():
        return
    items = sorted(((k, v) for k, v in aliases.items() if k and v and k != v), key=lambda kv: -len(kv[0]))
    skip = {'index.json', 'briefs_index.json', 'regions_index.json'}
    changed = 0
    for fp in out_dir.glob('*.json'):
        if fp.name in skip:
            continue
        try:
            d = json.loads(fp.read_text(encoding='utf-8'))
        except Exception:
            continue
        txt = d.get('text')
        if not isinstance(txt, str):
            continue
        new = txt
        for k, v in items:
            if k in new:
                new = new.replace(k, v)
        if new != txt:
            d['text'] = new
            fp.write_text(json.dumps(d, ensure_ascii=False), encoding='utf-8')
            changed += 1
    if changed:
        print(f"[OK] 브리핑 게임명 한글 치환: {changed}개 파일")


def collect_all_countries(countries=None, limit=100):
    """대상 국가별 iOS 매출·다운 차트 + 메타 + 한국어 게임명을 data/charts/{country}/{kind}/{date}.json 에 저장.
    기존 KR 파이프라인과 독립(추가 수집). 한 국가 실패해도 다음으로 진행."""
    countries = countries or COUNTRIES
    today = datetime.now().strftime('%Y-%m-%d')
    collected = {}
    for cc in countries:
        collected[cc] = {}
        for kind, slug in [('grossing', 'topgrossingapplications'), ('free', 'topfreeapplications')]:
            apps = fetch_apple_chart(cc, slug, limit)
            if not apps:
                continue
            try:
                attach_genres(apps, cc)
            except Exception as e:
                print(f"[WARN] {cc}/{kind} 장르 부착 실패: {e}")
            collected[cc][kind] = apps
            time.sleep(0.4)   # iTunes 레이트리밋 배려
    try:
        localize_titles(collected)          # title_kr 부여(KR→US→원본)
    except Exception as e:
        print(f"[WARN] 게임명 현지화 실패: {e}")
    ok = 0
    for cc in collected:
        for kind in collected[cc]:
            for a in collected[cc][kind]:
                a.setdefault('title_kr', a.get('title', ''))
            d = CHARTS_DIR / cc / kind
            d.mkdir(parents=True, exist_ok=True)
            (d / f'{today}.json').write_text(json.dumps(collected[cc][kind], ensure_ascii=False), encoding='utf-8')
            ok += 1
            print(f"[OK] {cc}/{kind}: {len(collected[cc][kind])}개 저장")
    print(f"[OK] 다국가 수집 완료: {ok}개(국가×차트) → {CHARTS_DIR}")
    return collected


CC_NAME_KR = {'kr': '한국', 'us': '미국', 'jp': '일본', 'cn': '중국', 'tw': '대만', 'gb': '영국',
              'de': '독일', 'fr': '프랑스', 'ca': '캐나다', 'au': '호주', 'it': '이탈리아', 'es': '스페인',
              'nl': '네덜란드', 'ru': '러시아', 'se': '스웨덴', 'sa': '사우디', 'ae': 'UAE', 'eg': '이집트',
              'tr': '터키', 'br': '브라질', 'mx': '멕시코', 'ar': '아르헨티나', 'co': '콜롬비아', 'id': '인도네시아',
              'th': '태국', 'vn': '베트남', 'ph': '필리핀', 'my': '말레이시아', 'sg': '싱가포르', 'in': '인도',
              'pk': '파키스탄', 'bd': '방글라데시'}

# 주요 시장 = 개별 탭(글로벌 모바일 매출 최상위). 순서 = 탭 노출 순서.
MAJOR_MARKETS = ['kr', 'us', 'jp', 'cn', 'tw']

# 지역 그룹(중·소규모만 묶음, 주요 시장 제외). 유럽은 한 그룹(러시아 포함).
REGIONS = {
    'europe': ('유럽', ['gb', 'de', 'fr', 'it', 'es', 'nl', 'se', 'ru']),
    'na_oceania': ('북미·오세아니아', ['ca', 'au']),
    'middle_east': ('중동', ['sa', 'ae', 'eg', 'tr']),
    'latam': ('중남미', ['br', 'mx', 'ar', 'co']),
    'sea': ('동남아', ['id', 'th', 'vn', 'ph', 'my', 'sg']),
    'south_asia': ('남아시아', ['in', 'pk', 'bd']),
}

# 브리핑 캐던스: 7일 이내 생성분이 있으면 재생성 스킵(주간 주기·비용 절감)
BRIEF_CADENCE_DAYS = 7


def _brief_fresh(fp, days=BRIEF_CADENCE_DAYS):
    """브리핑 파일이 days일 이내 생성됐으면 True(재생성 스킵)."""
    try:
        if not fp.exists():
            return False
        d = json.loads(fp.read_text(encoding='utf-8'))
        gen = (d.get('generated', '') or '')[:10]
        if not gen:
            return False
        last = datetime.strptime(gen, '%Y-%m-%d')
        return (datetime.now() - last).days < days
    except Exception:
        return False


def _brief_localized(fp):
    """브리핑이 한국어 게임명(title_kr) 기준으로 생성됐는지 마커 확인. 없으면 1회 강제 재생성용 False."""
    try:
        return bool(json.loads(fp.read_text(encoding='utf-8')).get('localized'))
    except Exception:
        return False


def _brief_structured(fp):
    """브리핑이 고정 항목(## 섹션) 구조인지 확인. 옛 무구조 브리핑은 False → 1회 강제 재생성."""
    try:
        return '## ' in (json.loads(fp.read_text(encoding='utf-8')).get('text') or '')
    except Exception:
        return False


def _brief_history(fp, cap=12):
    """재생성 시 직전 버전을 과거 기록(history)으로 보존. 현재 파일의 {generated,text}를 history 맨 앞에 추가(최대 cap)."""
    hist = []
    if fp.exists():
        try:
            old = json.loads(fp.read_text(encoding='utf-8'))
            hist = list(old.get('history') or [])
            if old.get('text') and old.get('generated'):
                hist.insert(0, {'date': (old.get('generated') or '')[:10], 'generated': old.get('generated'), 'text': old.get('text')})
        except Exception:
            pass
    # 같은 날짜 중복 제거(최신 우선)
    seen, out = set(), []
    for h in hist:
        d = h.get('date')
        if d in seen:
            continue
        seen.add(d); out.append(h)
    return out[:cap]


# ===== 시간축별(주/월/분기/년) 브리핑: '완료된 축'만 생성(미완성=데이터 부족→생성 안 함) =====
BRIEF_AXES = [('weekly', '주간'), ('monthly', '월간'), ('quarterly', '분기'), ('yearly', '연간')]
_AXIS_WIN = {'주간': '최근 주', '월간': '최근 달', '분기': '최근 분기', '연간': '최근 1년'}


def _market_axis_digest(market_key, axis_key):
    """commit된 docs/markets/{key}.json의 해당 시간축(완료구간) 윈도우 다이제스트. 완료구간 없으면 None(=미완성→생성 안 함)."""
    fp = Path('docs') / 'markets' / f'{market_key}.json'
    if not fp.exists():
        return None
    try:
        from collections import Counter
        mj = json.loads(fp.read_text(encoding='utf-8'))
        g = mj['charts']['grossing']
        tf = (g.get('timeframes') or {}).get(axis_key) or {}
        labels = tf.get('labels') or []
        if not labels:
            return None
        series = tf.get('series') or {}
        games = g.get('games') or {}
        li = len(labels) - 1
        rows = []
        for aid, s in series.items():
            now = s[li] if li < len(s) else None
            if now is None:
                continue
            prev = next((v for v in s if v is not None), None)
            meta = games.get(aid) or {}
            rows.append((now, prev, meta.get('title_kr') or meta.get('title') or aid, meta.get('genre', '')))
        if not rows:
            return None
        rows.sort(key=lambda r: r[0])
        top = ', '.join(f"{r[0]}.{r[2]}({r[3]})" for r in rows[:5])
        movers = [(r[2], r[1] - r[0]) for r in rows if r[1] is not None]
        ris = ', '.join(f"{m[0]}(▲{m[1]})" for m in sorted([m for m in movers if m[1] >= 1], key=lambda m: -m[1])[:3]) or '—'
        fal = ', '.join(f"{m[0]}(▼{abs(m[1])})" for m in sorted([m for m in movers if m[1] <= -1], key=lambda m: m[1])[:3]) or '—'
        gc = Counter((r[3] or '기타') for r in rows[:30])
        genres = ', '.join(f"{k} {v}" for k, v in gc.most_common(6))
        return f"기간 {labels[0]}~{labels[-1]}({len(labels)}구간) | 매출TOP5: {top} | 급상승: {ris} | 급하락: {fal} | 장르분포(상위30): {genres}"
    except Exception:
        return None


def _axes_prev(fp):
    """기존 파일의 axes. 구버전(단일 text)이면 weekly 축으로 이행(이력 보존)."""
    try:
        d = json.loads(fp.read_text(encoding='utf-8'))
    except Exception:
        return {}
    ax = dict(d.get('axes') or {})
    if 'weekly' not in ax and d.get('text'):
        ax['weekly'] = {'generated': d.get('generated'), 'text': d.get('text'), 'history': d.get('history') or []}
    return ax


def _axis_fresh(o, days=BRIEF_CADENCE_DAYS):
    try:
        gen = ((o or {}).get('generated', '') or '')[:10]
        if not gen:
            return False
        return (datetime.now() - datetime.strptime(gen, '%Y-%m-%d')).days < days
    except Exception:
        return False


def _axis_hist(prev, cap=12):
    hist = list((prev or {}).get('history') or [])
    t, g = (prev or {}).get('text'), (prev or {}).get('generated')
    if t and g:
        hist.insert(0, {'date': (g or '')[:10], 'generated': g, 'text': t})
    seen, out = set(), []
    for h in hist:
        d = h.get('date')
        if d in seen:
            continue
        seen.add(d)
        out.append(h)
    return out[:cap]


def _prior_signal(prev, cap=300):
    """직전 브리핑에서 검증할 가치가 있는 신호(움직임·PM 시사점·진출 전략·횡단 신호)만 짧게 추출.
    연속성 참조용·저비중 — 프롬프트에 '검증하되 반복 말라'로 주입. cite/링크 태그 정리·길이 캡."""
    import re as _re
    t = (prev or {}).get('text') or ''
    if not t or '## ' not in t:
        return ''
    picks = []
    for m in _re.finditer(r'##\s*(움직임|PM 시사점|진출 전략|횡단 신호)\s*\n(.+?)(?=\n##|\Z)', t, _re.S):
        body = _re.sub(r'\s+', ' ', _re.sub(r'<[^>]+>|\[([^\]]+)\]\([^)]+\)', r'\1', m.group(2))).strip()
        if body:
            picks.append(f"{m.group(1)}: {body}")
    s = ' / '.join(picks).strip()
    if not s:
        return ''
    when = ((prev or {}).get('generated') or '')[:10]
    return ((f"({when}) " if when else '') + s)[:cap]


AXIS_PV = 'v9-brief2'   # 브리핑 프롬프트 버전(바꾸면 캐시 무효화·전 시장 1회 재생성). v9: 두괄식·섹션명 실무화·짧은 문장·요약+세부·대시 금지

# [측정] API 사용량 집계 — 동작/비용 변화 없음, 기록만. 종료 시 atexit으로 총합·상위 비용 호출 출력.
import atexit as _atexit
USAGE_ROWS = []
def _print_usage_summary():
    if not USAGE_ROWS:
        return
    _ti = sum(r['in'] for r in USAGE_ROWS); _to = sum(r['out'] for r in USAGE_ROWS)
    _tcr = sum(r['cr'] for r in USAGE_ROWS); _tcc = sum(r['cc'] for r in USAGE_ROWS); _tws = sum(r['ws'] for r in USAGE_ROWS)
    print("=" * 64)
    print(f"[USAGE 총합] 호출 {len(USAGE_ROWS)}건 | 입력 {_ti:,} | 출력(사고포함) {_to:,} | 캐시읽기 {_tcr:,} | 캐시생성 {_tcc:,} | 웹검색 {_tws}건")
    for r in sorted(USAGE_ROWS, key=lambda x: x['out'], reverse=True)[:14]:
        print(f"   - {r['label']}: out={r['out']:,} in={r['in']:,} ws={r['ws']} {'[검색]' if r['search'] else ''}")
    print("=" * 64)
_atexit.register(_print_usage_summary)

# 모든 AI 인사이트 공통 — 현직 게임업계에서 통용되는 용어·관점·표현만 쓰도록 강제(업계인이 한눈에 이해).
INDUSTRY_VOICE = (
    "\n\n[작성 원칙 · 매우 중요] 현직 게임 사업/운영/PM 실무자가 한눈에 이해하는, 업계에서 실제로 쓰는 "
    "용어·관점·표현만 써라. 업계에서 안 쓰거나 지어낸 추상어는 금지(예: '운영형 전략' 같은 말은 실무자가 못 알아본다). "
    "추상적 분류·개념어를 만들지 말고, 무슨 일이 일어났고 PM이 무슨 판단을 내릴지를 구체적으로 써라. "
    "실무 어휘 예시(필요할 때만 자연스럽게, 억지 나열 금지): 라이브 서비스·라이브 운영, 콘텐츠 업데이트 주기, "
    "신작 모멘텀·안착, 흥행·롱런, 매출 순위 방어, BM(수익모델)·과금·가챠·배틀패스, 리텐션, "
    "UA(유저 확보)·마케팅 드라이브, 트래픽, IP·콜라보, 크로스프로모션, 점유율·경쟁 강도, 포화·틈새, 현지화·퍼블리싱. "
    "한 문장이라도 실무자가 '이게 무슨 말이지?' 하면 실패다."
)


GROUNDING = (
    "\n\n[근거·출처 · 매우 중요] 이번 기간에 눈에 띄게 급상승하거나 신규 진입한 게임이 있으면, "
    "신뢰할 수 있는 게임 전문 매체에서 그 원인(출시 초기 효과·대규모 프로모션·스트리머 마케팅·업데이트·이벤트·콜라보·IP 등)을 "
    "웹검색으로 확인해 한 줄로 덧붙여라. 확인되면 근거가 된 핵심 단어에 마크다운 링크 [단어](기사URL) 형태로 출처를 건다(URL을 따로 나열하지 말고 단어에만). "
    "규칙 — (1) 편집된 뉴스 기사만 출처로 인정. 커뮤니티·게시판·댓글·SNS·위키는 절대 금지. "
    "(2) 신뢰 출처로 확인되지 않으면 원인을 지어내지 말고 그냥 적지 마라(근거 없는 단정 금지). "
    "(3) 정황·추정 수준이면 '~로 보인다'로 헤지하되, 그것도 신뢰 매체 보도가 있을 때만. "
    "(4) 링크는 한 브리핑에서 꼭 필요한 2~4개로 절제한다. "
    "(5) 출처의 수치는 그대로 인용하고 반올림·확대 단정 금지 — 예: $811.9m을 '$1bn 돌파'로 올리지 마라. "
    "근사치는 '약 8억 달러'·'~수준'으로 헤지하고, 출처에 없는 수치는 만들지 마라."
)


def _axis_prompt(scope_label, axis_label, digest, is_region, grounded=False, prior_note=''):
    scope = f"'{scope_label}' 지역(여러 나라 합산) 시장" if is_region else f"'{scope_label}' 단일 시장"
    win = _AXIS_WIN.get(axis_label, axis_label)
    s_struct = "지역 전체에서 강세 장르·퍼블리셔 점유" if is_region else "강세 장르·퍼블리셔가 매출을 얼마나 쥐는지(점유)"
    _pb = (f"\n\n[직전 '{axis_label}' 분석 메모 — 저비중·연속성 검증용] {prior_note}\n"
           "→ 위 신호가 이번 데이터로도 유효한지 '## 움직임'에 한 줄로만 녹여라(맞으면 '지난 분석대로 ~', 바뀌었으면 '~로 전환'). "
           "근거 약하면 무시하고, 단순 반복·복붙·새 항목 추가는 금지. 이 메모는 참고일 뿐 현재 데이터가 우선이다.") if prior_note else ""
    return (f"다음은 App Store 게임 '매출' 차트의 '{scope_label}' {axis_label} 스냅샷·추이다.\n\n"
            f"[{scope_label} · {axis_label}] {digest}\n\n"
            f"게임 사업 PM이 {scope}을 '{axis_label}'({win}) 시간축 관점에서 읽도록, 아래 5개 항목을 정확히 이 순서·제목으로 써라. "
            "각 항목은 '## 제목' 한 줄로 시작하고, 그 아래 불릿(-)으로 쓴다. 각 항목 첫 불릿은 한 줄 핵심 요약(짧게), 이후 1~2개 불릿은 근거·세부. 항목 제목은 그대로 둘 것.\n"
            "## 시장 구조\n## 주요 게임·퍼블리셔\n## 주요 순위 변동\n## 장르 기회\n## 핵심 시사점\n\n"
            f"각 항목은 '{axis_label}'({win}) 시간축 기준으로 해석한다. 시장 구조: {s_struct}. "
            "주요 게임·퍼블리셔: 이 기간 매출 상위 게임의 성격과 강한 퍼블리셔. "
            f"주요 순위 변동: 이 기간({win}) 진입·급상승·급하락 위주(근거 약하면 '- 데이터 누적 중'). "
            "장르 기회: 경쟁 약한데 성과 나는 틈새 또는 포화 장르. 핵심 시사점: 진출·벤치마크·현지화 결론을 한 줄로(가장 중요). "
            "굵게(**)는 게임명·장르·퍼블리셔·국가명에만. 이모지·구분선(---) 금지. 문장에 'ㅡ' 대시(em·en 대시 포함) 쓰지 말고 마침표·쉼표·괄호·콜론으로. 문장은 짧게(불릿 1개=1문장). 모호한 조어 금지, 실무에서 쓰는 명확한 용어로. 한국어, 군더더기 없이. '게임/Games/游戏'는 분석 카테고리이지 장르가 아니다(장르는 구체 서브장르 MMORPG·4X·매치3 등으로만, '상위 100이 전부 게임' 류 자명한 서술 금지). 게임명 고유명사 외엔 한국어로만(游戏·ゲーム 등 외국어 일반어 금지)." + _pb + INDUSTRY_VOICE + (GROUNDING if grounded else ""))


def _build_scope_axes(fp, scope_label, market_key, weekly_digest, is_region):
    """완료된 시간축만 브리핑 생성. weekly=신선한 스냅샷 digest, 그 외=commit된 market JSON 윈도우. 축별 7일 게이팅·이력 보존."""
    prev_axes = _axes_prev(fp)
    now_s = datetime.now().strftime('%Y-%m-%d %H:%M')
    axes = {}
    for axis_key, axis_label in BRIEF_AXES:
        digest = weekly_digest if axis_key == 'weekly' else _market_axis_digest(market_key, axis_key)
        if axis_key == 'weekly' and digest:   # [데이터 보강] 전주 대비 순위변동(WoW)을 커밋된 주간 시계열에서 덧붙여 '움직임' 근거 강화(없으면 그대로)
            _mv = _market_axis_digest(market_key, 'weekly')
            if _mv and '급상승:' in _mv:
                _seg = _mv.split('장르분포')[0]; _i = _seg.find('급상승:')
                if _i >= 0:
                    digest = digest + ' | (전주대비) ' + _seg[_i:].strip().rstrip('| ').strip()
        if not digest:
            if axis_key in prev_axes:   # 미완성이지만 과거 생성분 있으면 유지
                axes[axis_key] = prev_axes[axis_key]
            continue
        prev = prev_axes.get(axis_key) or {}
        if _axis_fresh(prev) and '## ' in (prev.get('text') or '') and prev.get('pv') == AXIS_PV:
            axes[axis_key] = prev       # 7일 이내 최신 + 같은 프롬프트버전 → 재사용(비용 절감)
            continue
        try:
            _grounded = (axis_key == 'weekly')   # #11: 최근(주간) 축만 웹검색으로 원인·출처 보강
            text = call_claude_with_retry(_axis_prompt(scope_label, axis_label, digest, is_region, grounded=_grounded, prior_note=_prior_signal(prev)), max_tokens=MAX_OUTPUT_TOKENS, web_search=_grounded, usage_label=f"{scope_label}/{axis_label}")
        except Exception as e:
            print(f'[WARN] {scope_label} {axis_label} 브리핑 실패: {e}')
            if axis_key in prev_axes:
                axes[axis_key] = prev_axes[axis_key]
            continue
        axes[axis_key] = {'generated': now_s, 'text': text, 'pv': AXIS_PV, 'history': _axis_hist(prev)}
        print(f'[OK] {scope_label} · {axis_label} 브리핑 생성')
    return axes


def _country_digest(g):
    """국가 매출 차트 1개 → 'TOP5 + 장르분포' 한 줄 다이제스트."""
    from collections import Counter
    top = ', '.join(f"{a.get('rank')}.{a.get('title_kr') or a.get('title')}({a.get('genre', '')})" for a in g[:5])
    gc = Counter(a.get('genre', '기타') for a in g)
    genres = ', '.join(f"{k} {v}" for k, v in gc.most_common(6))
    return f"매출TOP5: {top} | 장르분포(상위100): {genres}"


def build_major_briefs(collected):
    """주요 시장(개별 국가) 시간축별 브리핑 생성(완료된 축만, 축별 7일 게이팅). docs/markets/major_{cc}.json. 노출 가능한 [{key,name}] 반환."""
    if not collected:
        return []
    out_dir = Path('docs') / 'markets'
    out_dir.mkdir(parents=True, exist_ok=True)
    available = []
    for cc in MAJOR_MARKETS:
        g = (collected.get(cc, {}) or {}).get('grossing')
        if not g:
            continue  # 아직 수집 안 된 주요국(탭 미노출)
        name = CC_NAME_KR.get(cc, cc.upper())
        fp = out_dir / f'major_{cc}.json'
        axes = _build_scope_axes(fp, name, cc, _country_digest(g), is_region=False)
        if not axes:
            if fp.exists():
                available.append({'key': cc, 'name': name})
            continue
        wk = axes.get('weekly') or next((axes[k] for k, _ in BRIEF_AXES if k in axes), {})
        fp.write_text(json.dumps({'generated': wk.get('generated') or datetime.now().strftime('%Y-%m-%d %H:%M'),
                                  'market': name, 'cc': cc, 'localized': True, 'axes': axes,
                                  'text': wk.get('text') or '', 'history': wk.get('history') or []},
                                 ensure_ascii=False), encoding='utf-8')
        available.append({'key': cc, 'name': name})
        print(f'[OK] 주요시장 브리핑 저장: major_{cc}.json ({name}, 축 {list(axes.keys())})')
    return available


def write_briefs_index(majors, regions):
    """탭 노출 순서: 주요시장(개별) → 지역. docs/markets/briefs_index.json."""
    out_dir = Path('docs') / 'markets'
    out_dir.mkdir(parents=True, exist_ok=True)
    tabs = [{'type': 'major', 'key': m['key'], 'name': m['name']} for m in (majors or [])]
    tabs += [{'type': 'region', 'key': r['key'], 'name': r['name'], 'countries': r.get('countries')}
             for r in (regions or [])]
    (out_dir / 'briefs_index.json').write_text(
        json.dumps({'generated': datetime.now().strftime('%Y-%m-%d %H:%M'), 'tabs': tabs},
                   ensure_ascii=False), encoding='utf-8')
    print(f'[OK] 브리핑 인덱스 저장: briefs_index.json (주요 {len(majors or [])} + 지역 {len(regions or [])})')


def build_regional_briefs(collected):
    """중·소규모 지역 그룹별 브리핑 생성(Opus, 주간 게이팅). docs/markets/region_{key}.json. 노출 가능한 [{key,name,countries}] 반환."""
    if not collected:
        print('[INFO] 지역 브리핑 스킵(수집 데이터 없음)'); return []
    out_dir = Path('docs') / 'markets'
    out_dir.mkdir(parents=True, exist_ok=True)
    available = []
    for key, (name, ccs) in REGIONS.items():
        members = [cc for cc in ccs if (collected.get(cc, {}) or {}).get('grossing')]
        fp = out_dir / f'region_{key}.json'
        if not members:
            # 이번 수집엔 멤버 없지만 기존 생성분이 있으면 탭 유지(수동 32개국 확장 스냅샷 보존)
            if fp.exists():
                try:
                    prev = json.loads(fp.read_text(encoding='utf-8'))
                    available.append({'key': key, 'name': name, 'countries': prev.get('countries')})
                except Exception:
                    pass
            continue
        weekly_digest = '\n'.join(f"[{CC_NAME_KR.get(cc, cc.upper())}] {_country_digest(collected[cc]['grossing'])}" for cc in members)
        axes = _build_scope_axes(fp, name, key, weekly_digest, is_region=True)
        if not axes:
            available.append({'key': key, 'name': name, 'countries': len(members)})
            continue
        wk = axes.get('weekly') or next((axes[k] for k, _ in BRIEF_AXES if k in axes), {})
        fp.write_text(json.dumps({'generated': wk.get('generated') or datetime.now().strftime('%Y-%m-%d %H:%M'),
                                  'region': name, 'countries': len(members), 'localized': True, 'axes': axes,
                                  'text': wk.get('text') or '', 'history': wk.get('history') or []},
                                 ensure_ascii=False), encoding='utf-8')
        available.append({'key': key, 'name': name, 'countries': len(members)})
        print(f'[OK] 지역 브리핑 저장: region_{key}.json ({name}, {len(members)}개국, 축 {list(axes.keys())})')
    return available


def build_global_brief(collected):
    """주요시장(개별)+지역 브리핑을 토대로 전 시장 횡단 글로벌 헤드라인을 종합 생성(Opus). docs/markets/global_brief.json.
    계층적: major_*.json + region_*.json을 입력으로 합성 — 둘 다 없으면 국가 다이제스트로 폴백."""
    if not collected:
        print('[INFO] 글로벌 브리핑 스킵(수집 데이터 없음)'); return
    out_dir = Path('docs') / 'markets'
    out_dir.mkdir(parents=True, exist_ok=True)
    gp = out_dir / 'global_brief.json'

    def _read(p):
        try:
            return json.loads(p.read_text(encoding='utf-8'))
        except Exception:
            return None

    n_countries = sum(1 for cc, k in collected.items() if (k or {}).get('grossing'))
    prev_axes = _axes_prev(gp)
    now_s = datetime.now().strftime('%Y-%m-%d %H:%M')
    g_axes = {}
    total_sources = 0
    for axis_key, axis_label in BRIEF_AXES:
        # 같은 시간축의 하위 브리핑(주요·지역) 수집 — 그 축이 있을 때만
        subs, newest_sub = [], ''
        for cc in MAJOR_MARKETS:
            d = _read(out_dir / f'major_{cc}.json')
            ax = ((d or {}).get('axes') or {}).get(axis_key) if d else None
            if ax and ax.get('text'):
                subs.append((d.get('market', CC_NAME_KR.get(cc, cc)) + ' (주요 단일 시장)', ax['text']))
                newest_sub = max(newest_sub, ax.get('generated', '') or '')
        for key, (name, _ccs) in REGIONS.items():
            d = _read(out_dir / f'region_{key}.json')
            ax = ((d or {}).get('axes') or {}).get(axis_key) if d else None
            if ax and ax.get('text'):
                subs.append((d.get('region', name) + ' (지역)', ax['text']))
                newest_sub = max(newest_sub, ax.get('generated', '') or '')
        if not subs:
            if axis_key in prev_axes:   # 이 축 하위 분석이 아직 없으면 과거 종합 유지
                g_axes[axis_key] = prev_axes[axis_key]
            continue
        total_sources = max(total_sources, len(subs))
        prev = prev_axes.get(axis_key) or {}
        subs_newer = bool(newest_sub) and newest_sub > ((prev.get('generated') or '')[:16])
        if _axis_fresh(prev) and '## ' in (prev.get('text') or '') and not subs_newer and prev.get('pv') == AXIS_PV:
            g_axes[axis_key] = prev
            continue
        win = _AXIS_WIN.get(axis_label, axis_label)
        body = '\n\n'.join(f"# {label} 분석\n{tx}" for label, tx in subs)
        _gp = _prior_signal(prev)
        _gpb = (f"\n\n(직전 '{axis_label}' 글로벌 메모(저비중·연속성 검증용) {_gp}\n→ 이번 종합과 비교해 유효/전환 여부만 '## 전 시장 공통 흐름'에 한 줄로 녹여라(반복·복붙·새 항목 금지). 현재 분석이 우선이다.") if _gp else ""
        prompt = (f"다음은 같은 기간 App Store 게임 매출 차트를, 주요 시장은 개별·중소규모는 지역으로 묶어 '{axis_label}'({win}) 시간축으로 분석한 결과다.\n\n{body}\n\n"
                  f"위 '{axis_label}' 분석들을 토대로 종합하여, 게임 사업 PM이 전 세계를 횡단해 읽을 '{axis_label}'({win}) 헤드라인을 아래 4개 항목으로 써라. "
                  "각 항목은 '## 제목' 한 줄로 시작하고, 그 아래 불릿(-)으로 쓴다. 각 항목 첫 불릿은 한 줄 핵심 요약(짧게), 이후 1~2개 불릿은 근거·세부. 항목 제목은 그대로 둘 것.\n"
                  "## 전 시장 공통 흐름\n## 시장별 차이\n## IP·퍼블리셔\n## 진출 전략\n\n"
                  "각 항목 내용 가이드. 전 시장 공통 흐름: 여러 시장·지역에서 동시에 강한 게임·장르. "
                  "시장별 차이: 주요 시장 간·지역 간 장르 구성 대비. IP·퍼블리셔: 여러 시장을 관통하는 글로벌 IP·퍼블리셔. "
                  "진출 전략: 다음 진출·벤치마크 시장 결론(가장 중요). "
                  "하위 분석에 없는 사실을 지어내지 말 것. 굵게(**)는 게임명·장르·퍼블리셔·국가/지역명에만. 이모지·구분선(---) 금지. 문장에 'ㅡ' 대시(em·en 대시 포함) 쓰지 말고 마침표·쉼표·괄호·콜론으로. 문장은 짧게(불릿 1개=1문장). 모호한 조어 금지, 실무에서 쓰는 명확한 용어로. 한국어, 군더더기 없이. '게임/Games/游戏'는 분석 카테고리이지 장르가 아니다(장르는 구체 서브장르 MMORPG·4X·매치3 등으로만, '상위 100이 전부 게임' 류 자명한 서술 금지). 게임명 고유명사 외엔 한국어로만(游戏·ゲーム 등 외국어 일반어 금지)." + _gpb + INDUSTRY_VOICE + (GROUNDING if axis_key == 'weekly' else ""))
        try:
            text = call_claude_with_retry(prompt, max_tokens=MAX_OUTPUT_TOKENS, web_search=(axis_key == 'weekly'), usage_label=f"글로벌/{axis_label}")
        except Exception as e:
            print(f'[WARN] 글로벌 {axis_label} 브리핑 생성 실패: {e}')
            if axis_key in prev_axes:
                g_axes[axis_key] = prev_axes[axis_key]
            continue
        g_axes[axis_key] = {'generated': now_s, 'text': text, 'pv': AXIS_PV, 'history': _axis_hist(prev)}
        print(f'[OK] 글로벌 · {axis_label} 브리핑 생성(sources={len(subs)})')
    if not g_axes:
        print('[INFO] 글로벌 브리핑 스킵(하위 분석 없음)'); return
    wk = g_axes.get('weekly') or next((g_axes[k] for k, _ in BRIEF_AXES if k in g_axes), {})
    gp.write_text(json.dumps({'generated': wk.get('generated') or now_s,
                              'countries': n_countries, 'basis': 'hierarchical', 'sources': total_sources,
                              'localized': True, 'axes': g_axes,
                              'text': wk.get('text') or '', 'history': wk.get('history') or []},
                             ensure_ascii=False), encoding='utf-8')
    print(f'[OK] 글로벌 브리핑 저장: docs/markets/global_brief.json (축 {list(g_axes.keys())}, sources={total_sources})')


def load_data_in_date_range(start_dt, end_dt, today_dt=None, current=None):
    """[start_dt, end_dt] 범위 데이터 로드. today_dt와 current 주어지면 그날은 current로 대체."""
    result = []
    d = start_dt
    while d.date() <= end_dt.date():
        if today_dt is not None and d.date() == today_dt.date():
            if current is not None:
                result.append(current)
        else:
            data = load_data_by_date(d.strftime('%Y-%m-%d'))
            if data is not None:
                result.append(data)
        d += timedelta(days=1)
    return result


# ============================================================
# 2. 분기 유틸
# ============================================================

def get_quarter(date_dt):
    return date_dt.year, (date_dt.month - 1) // 3 + 1


def get_quarter_range(year, quarter):
    start_month = (quarter - 1) * 3 + 1
    start = datetime(year, start_month, 1)
    if quarter == 4:
        end = datetime(year, 12, 31, 23, 59, 59)
    else:
        end = datetime(year, start_month + 3, 1) - timedelta(seconds=1)
    return start, end


def get_prior_quarter(year, quarter):
    if quarter == 1:
        return year - 1, 4
    return year, quarter - 1


# ============================================================
# 3. 평균 순위 + 비교 계산
# ============================================================

def compute_average_ranks(window_data_list):
    """N일치 데이터에서 게임별 평균 순위. 빈 리스트면 빈 dict."""
    if not window_data_list:
        return {}
    accumulator = {}
    for day_data in window_data_list:
        for app in day_data:
            app_id = app['app_id']
            if not app_id:
                continue
            if app_id not in accumulator:
                accumulator[app_id] = {'rank_sum': 0, 'days': 0,
                                       'title': app['title'], 'developer': app['developer']}
            accumulator[app_id]['rank_sum'] += app['rank']
            accumulator[app_id]['days'] += 1
    total_days = len(window_data_list)
    return {
        app_id: {
            'avg_rank': round(info['rank_sum'] / info['days'], 1),
            'title': info['title'],
            'developer': info['developer'],
            'days_in_chart': info['days'],
            'total_days': total_days,
        }
        for app_id, info in accumulator.items()
    }


def _surge_threshold(prev_rank, curr_rank, base):
    """급변 기준 계단 = 기간 기본값(base) × 위치 보정. 상위권일수록 민감, 하위권은 둔감."""
    best = min(prev_rank, curr_rank)
    if best <= 10:
        m = 0.5
    elif best <= 30:
        m = 0.75
    elif best <= 60:
        m = 1.0
    else:
        m = 1.5
    return base * m


def compute_simple_changes(previous, current, threshold=10):
    """1일선용 단순 두 시점 비교."""
    if not previous:
        return {'mode': 'simple', 'new_entries': [], 'dropped': [], 'rank_changes': [],
                'past_days': 0, 'recent_days': 1, 'comparable': False}

    prev_by_id = {item['app_id']: item for item in previous if item['app_id']}
    curr_by_id = {item['app_id']: item for item in current if item['app_id']}

    new_entries = [
        {'title': c['title'], 'developer': c['developer'], 'rank': c['rank']}
        for c in current if c['app_id'] and c['app_id'] not in prev_by_id
    ]
    dropped = [
        {'title': p['title'], 'developer': p['developer'], 'rank': p['rank']}
        for p in previous if p['app_id'] and p['app_id'] not in curr_by_id
    ]
    rank_changes = []
    for app_id, curr in curr_by_id.items():
        if app_id in prev_by_id:
            prev_rank = prev_by_id[app_id]['rank']
            curr_rank = curr['rank']
            diff = prev_rank - curr_rank
            if abs(diff) >= _surge_threshold(prev_rank, curr_rank, threshold):
                rank_changes.append({
                    'title': curr['title'], 'developer': curr['developer'],
                    'prev_rank': prev_rank, 'curr_rank': curr_rank, 'change': diff,
                })
    return {'mode': 'simple', 'new_entries': new_entries, 'dropped': dropped,
            'rank_changes': rank_changes, 'past_days': 1, 'recent_days': 1, 'comparable': True}


def compute_period_changes(past_data, recent_data, threshold=5):
    """이동평균 기반 두 기간 비교. 한쪽이라도 비면 빈 결과 반환."""
    past_avg = compute_average_ranks(past_data)
    recent_avg = compute_average_ranks(recent_data)

    base = {
        'mode': 'moving_average',
        'past_days': len(past_data),
        'recent_days': len(recent_data),
    }

    if not past_avg or not recent_avg:
        base.update({'new_entries': [], 'dropped': [], 'rank_changes': [], 'comparable': False})
        return base

    new_entries = [
        {'title': info['title'], 'developer': info['developer'],
         'avg_rank': round(info['avg_rank'], 1),
         'days_in_chart': info['days_in_chart'], 'total_days': info['total_days']}
        for app_id, info in recent_avg.items() if app_id not in past_avg
    ]
    dropped = [
        {'title': info['title'], 'developer': info['developer'],
         'avg_rank': round(info['avg_rank'], 1),
         'days_in_chart': info['days_in_chart'], 'total_days': info['total_days']}
        for app_id, info in past_avg.items() if app_id not in recent_avg
    ]
    rank_changes = []
    for app_id, recent_info in recent_avg.items():
        if app_id in past_avg:
            past_r = past_avg[app_id]['avg_rank']
            recent_r = recent_info['avg_rank']
            diff = past_r - recent_r
            if abs(diff) >= _surge_threshold(past_r, recent_r, threshold):
                rank_changes.append({
                    'title': recent_info['title'], 'developer': recent_info['developer'],
                    'prev_rank': round(past_r, 1), 'curr_rank': round(recent_r, 1),
                    'change': round(diff, 1),
                    'recent_days': recent_info['days_in_chart'],
                    'prior_days': past_avg[app_id]['days_in_chart'],
                })
    base.update({'new_entries': new_entries, 'dropped': dropped, 'rank_changes': rank_changes,
                 'comparable': True})
    return base


# ============================================================
# 4. 유의문구 생성
# ============================================================

def generate_warning(changes, expected_past, expected_recent, past_label, recent_label):
    past_d = changes.get('past_days', 0)
    recent_d = changes.get('recent_days', 0)

    if past_d == 0 and recent_d == 0:
        return f"⚠️ 양쪽 기간 모두 데이터 없음 (봇 시작 이전 또는 누락). 비교 불가."
    if past_d == 0:
        return f"⚠️ {past_label} 데이터 없음. 비교 기준 부재 — 변화 분석 불가, {recent_label} 평균만 의미 있음."
    if recent_d == 0:
        return f"⚠️ {recent_label} 데이터 없음. 비교 불가."

    warnings = []
    if past_d < expected_past:
        warnings.append(f"{past_label}: {past_d}/{expected_past}일")
    if recent_d < expected_recent:
        warnings.append(f"{recent_label}: {recent_d}/{expected_recent}일")
    if warnings:
        return f"⚠️ 부분 데이터: {' / '.join(warnings)}"
    return None


# ============================================================
# 5. 시간축별 분석 함수
# ============================================================

def analyze_daily(today_dt, current):
    """1일선: 어제 vs 오늘."""
    previous, prev_date = find_most_recent_past_data()
    changes = compute_simple_changes(previous, current, threshold=8)
    changes['past_label'] = prev_date if prev_date else '어제 (데이터 없음)'
    changes['recent_label'] = today_dt.strftime('%Y-%m-%d')
    changes['period_label'] = f"{changes['past_label']} → {changes['recent_label']}"
    if not changes['comparable']:
        changes['warning'] = "⚠️ 어제 데이터 없음. 비교 불가 (첫 실행이거나 누락)."
    else:
        changes['warning'] = None
    changes['period_key'] = today_dt.strftime('%Y-%m-%d')  # 매일 바뀜 → 일별 인사이트 매일 갱신
    return changes


def analyze_weekly(today_dt, current):
    """1주선: 마지막으로 완료된 주(월~일) vs 그 전 주. 매일 호출해도 그 주 동안 결과 고정."""
    this_monday = today_dt - timedelta(days=today_dt.weekday())
    recent_start = this_monday - timedelta(days=7)   # 지난주 월요일
    recent_end = this_monday - timedelta(days=1)      # 지난주 일요일
    past_start = this_monday - timedelta(days=14)
    past_end = this_monday - timedelta(days=8)

    past_data = load_data_in_date_range(past_start, past_end)
    recent_data = load_data_in_date_range(recent_start, recent_end)

    changes = compute_period_changes(past_data, recent_data, threshold=5)
    past_label = f"{past_start.strftime('%Y-%m-%d')}~{past_end.strftime('%m-%d')} (전전주)"
    recent_label = f"{recent_start.strftime('%Y-%m-%d')}~{recent_end.strftime('%m-%d')} (전주)"
    changes['past_label'] = past_label
    changes['recent_label'] = recent_label
    changes['period_label'] = f"{past_label} 평균 → {recent_label} 평균"
    changes['warning'] = generate_warning(changes, 7, 7, '전전주', '전주')
    changes['period_key'] = recent_start.strftime('%Y-%m-%d')  # 주가 바뀔 때만 변경
    return changes


def analyze_monthly(today_dt, current):
    """1달선: 마지막으로 완료된 월 vs 그 전 월. 매일 호출해도 그 달 동안 결과 고정."""
    first_this_month = datetime(today_dt.year, today_dt.month, 1)
    last_month_end = first_this_month - timedelta(days=1)
    last_month_start = datetime(last_month_end.year, last_month_end.month, 1)
    two_months_end = last_month_start - timedelta(days=1)
    two_months_start = datetime(two_months_end.year, two_months_end.month, 1)

    past_data = load_data_in_date_range(two_months_start, two_months_end)
    recent_data = load_data_in_date_range(last_month_start, last_month_end)

    expected_past = (two_months_end.date() - two_months_start.date()).days + 1
    expected_recent = (last_month_end.date() - last_month_start.date()).days + 1

    changes = compute_period_changes(past_data, recent_data, threshold=5)
    past_label = f"{two_months_start.strftime('%Y-%m')} (전전월, {expected_past}일)"
    recent_label = f"{last_month_start.strftime('%Y-%m')} (전월, {expected_recent}일)"
    changes['past_label'] = past_label
    changes['recent_label'] = recent_label
    changes['period_label'] = f"{past_label} 평균 → {recent_label} 평균"
    changes['warning'] = generate_warning(changes, expected_past, expected_recent, '전전월', '전월')
    changes['period_key'] = last_month_start.strftime('%Y-%m')  # 달이 바뀔 때만 변경
    return changes


def analyze_quarterly(today_dt, current):
    """분기선: 전전분기 vs 전분기. 분기 시작일에 호출."""
    curr_year, curr_q = get_quarter(today_dt)
    prior_year, prior_q = get_prior_quarter(curr_year, curr_q)
    two_prior_year, two_prior_q = get_prior_quarter(prior_year, prior_q)

    prior_q_start, prior_q_end = get_quarter_range(prior_year, prior_q)
    two_prior_q_start, two_prior_q_end = get_quarter_range(two_prior_year, two_prior_q)

    past_data = load_data_in_date_range(two_prior_q_start, two_prior_q_end)
    recent_data = load_data_in_date_range(prior_q_start, prior_q_end)

    expected_past = (two_prior_q_end.date() - two_prior_q_start.date()).days + 1
    expected_recent = (prior_q_end.date() - prior_q_start.date()).days + 1

    changes = compute_period_changes(past_data, recent_data, threshold=5)
    past_label = f"{two_prior_year} Q{two_prior_q} (전전분기, {expected_past}일)"
    recent_label = f"{prior_year} Q{prior_q} (전분기, {expected_recent}일)"
    changes['past_label'] = past_label
    changes['recent_label'] = recent_label
    changes['period_label'] = f"{past_label} 평균 → {recent_label} 평균"
    changes['warning'] = generate_warning(changes, expected_past, expected_recent, '전전분기', '전분기')
    changes['period_key'] = f"{prior_year}-Q{prior_q}"  # 분기가 바뀔 때만 변경
    return changes


def analyze_yearly(today_dt, current):
    """1년선: 재작년 vs 작년. 1/1에 호출."""
    curr_year = today_dt.year
    prior_year = curr_year - 1
    two_prior_year = curr_year - 2

    prior_start = datetime(prior_year, 1, 1)
    prior_end = datetime(prior_year, 12, 31, 23, 59, 59)
    two_prior_start = datetime(two_prior_year, 1, 1)
    two_prior_end = datetime(two_prior_year, 12, 31, 23, 59, 59)

    past_data = load_data_in_date_range(two_prior_start, two_prior_end)
    recent_data = load_data_in_date_range(prior_start, prior_end)

    expected_past = (two_prior_end.date() - two_prior_start.date()).days + 1
    expected_recent = (prior_end.date() - prior_start.date()).days + 1

    changes = compute_period_changes(past_data, recent_data, threshold=5)
    past_label = f"{two_prior_year}년 (재작년, {expected_past}일)"
    recent_label = f"{prior_year}년 (작년, {expected_recent}일)"
    changes['past_label'] = past_label
    changes['recent_label'] = recent_label
    changes['period_label'] = f"{past_label} 평균 → {recent_label} 평균"
    changes['warning'] = generate_warning(changes, expected_past, expected_recent, '재작년', '작년')
    changes['period_key'] = str(prior_year)  # 해가 바뀔 때만 변경
    return changes


# ============================================================
# 6. 활성 시간축 판단
# ============================================================

def get_active_timeframes(today_dt):
    """모든 시간축을 매일 계산(표시 여부는 데이터 유무로 main에서 판단)."""
    return ['1일', '1주', '1달', '분기', '1년']


def run_active_analyses(today_dt, current, active_names):
    """활성 시간축 분석 실행."""
    analyzers = {
        '1일': analyze_daily, '1주': analyze_weekly, '1달': analyze_monthly,
        '분기': analyze_quarterly, '1년': analyze_yearly,
    }
    return {name: analyzers[name](today_dt, current) for name in active_names}


# ============================================================
# 7. Claude 요약 (재시도 로직 포함)
# ============================================================

def _filter_sources(text):
    """#11: 브리핑 본문의 마크다운 링크 [라벨](url) 중 신뢰 출처가 아니면 링크만 제거(라벨 텍스트는 유지)."""
    if not text:
        return text
    import re
    return re.sub(r'\[([^\]]+)\]\((https?://[^)\s]+)\)',
                  lambda m: m.group(0) if is_trusted_source(m.group(2)) else m.group(1),
                  text)


def call_claude_with_retry(prompt, max_tokens=MAX_OUTPUT_TOKENS, max_retries=4, web_search=False, usage_label=''):
    """Claude API 호출 (Opus 4.8 적응형 사고 + effort 최대).
    응답은 thinking 블록 뒤 text 블록 → text만 추출. 일시 오류 시 지수 백오프 재시도.
    web_search=True면 신뢰 도메인(TRUSTED_DOMAINS) 안에서만 웹검색 도구 사용 + 비신뢰 링크 후처리 제거.
    마지막 시도는 사고 옵션을 빼고(파라미터 문제 대비 안전망) 호출. 모두 실패 시 None."""
    client = Anthropic(api_key=ANTHROPIC_API_KEY)
    use_allowed = True  # 신뢰도메인 화이트리스트로 검색 제한; 일부가 크롤러 차단(400)이면 False로 떨궈 무제한+사후필터 폴백
    for attempt in range(1, max_retries + 1):
        try:
            kwargs = dict(
                model=CLAUDE_MODEL,
                max_tokens=max_tokens,
                messages=[{'role': 'user', 'content': prompt}],
            )
            if web_search and attempt < max_retries:  # #11: 신뢰 도메인 안에서 검색(마지막 시도는 안전망으로 검색 끔)
                _tool = {'type': 'web_search_20250305', 'name': 'web_search', 'max_uses': 2}
                if use_allowed:  # 크롤러 비접근 도메인으로 400 나면 use_allowed=False → 무제한 검색+사후필터
                    _tool['allowed_domains'] = TRUSTED_DOMAINS
                kwargs['tools'] = [_tool]
            if attempt < max_retries:  # 마지막 시도 전까지는 적응형 사고 + 최대 effort
                kwargs['thinking'] = {'type': 'adaptive'}
                kwargs['output_config'] = {'effort': THINKING_EFFORT}
            # 스트리밍 필수: max effort + 큰 max_tokens는 10분 초과 가능 → stream 사용
            with client.messages.stream(**kwargs) as stream:
                final = stream.get_final_message()
            try:  # [측정] 사용량 기록: 입력/출력(사고포함)/캐시/웹검색 횟수
                _u = getattr(final, 'usage', None)
                if _u is not None:
                    _it = getattr(_u, 'input_tokens', 0) or 0; _ot = getattr(_u, 'output_tokens', 0) or 0
                    _cr = getattr(_u, 'cache_read_input_tokens', 0) or 0; _cc = getattr(_u, 'cache_creation_input_tokens', 0) or 0
                    _stu = getattr(_u, 'server_tool_use', None); _ws = (getattr(_stu, 'web_search_requests', 0) or 0) if _stu else 0
                    USAGE_ROWS.append({'label': usage_label or '(call)', 'in': _it, 'out': _ot, 'cr': _cr, 'cc': _cc, 'ws': _ws, 'search': bool(web_search)})
                    print(f"[USAGE] {usage_label or '(call)'}: in={_it:,} out={_ot:,} cache_r={_cr:,} web_search={_ws}")
            except Exception as _ue:
                print(f"[USAGE] 사용량 기록 실패: {_ue}")
            # 웹검색 시 검색 전후로 text 블록이 여러 개 → 모두 이어붙임
            parts = [b.text for b in final.content
                     if getattr(b, 'type', None) == 'text' and getattr(b, 'text', None)]
            text = '\n'.join(parts).strip() if parts else None
            if text and web_search and '##' in text:  # 웹검색 준비멘트(검색 전 영어/한국어 혼잣말)가 앞에 붙음 → 본문 '##'부터만 남김
                text = text[text.index('##'):].strip()
            if text and web_search:  # 네이티브 인용 태그 <cite index="x-y">…</cite> 제거(렌더 깨짐 방지). 본문 텍스트는 유지, 출처는 [라벨](url) 마크다운 링크로.
                import re as _re
                text = _re.sub(r'</?cite[^>]*>', '', text)
            if text:
                return _filter_sources(text) if web_search else text
            print(f"[WARN] 응답에 text 블록 없음 (시도 {attempt}/{max_retries}) — 재시도")
        except Exception as e:
            if web_search and use_allowed and 'not accessible to our user agent' in str(e):
                use_allowed = False  # 신뢰도메인 일부가 크롤러 비접근 → allowed_domains 제거 후 무제한 검색+사후필터로 폴백
                print(f"[WARN] 일부 신뢰도메인 크롤러 비접근 → 무제한 검색+사후필터 폴백 후 재시도")
                continue
            print(f"[WARN] Claude API 실패 (시도 {attempt}/{max_retries}): {e}")
        if attempt < max_retries:
            wait = 2 ** attempt  # 2, 4, 8초
            print(f"       {wait}초 후 재시도...")
            time.sleep(wait)
        else:
            print("[ERROR] Claude API 재시도 모두 실패. 요약 없이 진행.")
            return None
    return None


def generate_daily_summary(current, changes, chart_used):
    """1일선만 활성인 평일용 짧은 요약."""
    if not changes['comparable']:
        return (f"비교 가능한 어제 데이터가 없습니다 (차트: {chart_used}). "
                f"다음 실행부터 1일 변동 분석이 시작됩니다.")

    prompt = f"""한국 App Store 게임 {chart_used} 차트의 1일 변동을 사업PM 관점에서 짧게.

[비교 구간] {changes['period_label']}

[신규 진입]
{json.dumps(changes.get('new_entries', [])[:10], ensure_ascii=False, indent=2)}

[차트 이탈]
{json.dumps(changes.get('dropped', [])[:10], ensure_ascii=False, indent=2)}

[큰 변동 (10등 이상)]
{json.dumps(changes.get('rank_changes', [])[:15], ensure_ascii=False, indent=2)}

한국어 2~3문단, 각 2~3줄:
1. 가장 주목할 변동 1~2건 (이유 추정 가능하면)
2. 사업PM 한 줄 인사이트

일간은 노이즈 많으니 진짜 시그널만. 미미하면 "특이사항 없음"이라고 솔직하게.""" + INDUSTRY_VOICE

    result = call_claude_with_retry(prompt, max_tokens=800)
    if result is None:
        return ("⚠️ AI 요약 생성 실패 — Claude API 일시적 과부하(529). "
                "차트 데이터와 변화 분석은 정상이며, 잠시 후 재실행하면 요약도 생성됩니다.")
    return result


def generate_comprehensive_summary(current, analyses, chart_used):
    """다중 시간축 종합 분석."""
    sections = []
    for name, ch in analyses.items():
        mode = "단순 비교" if ch.get('mode') == 'simple' else "이동평균"
        warning = ch.get('warning') or ""
        if not ch.get('comparable'):
            sections.append(f"\n[{name}선] ({mode}) {ch['period_label']}\n{warning}\n→ 비교 불가\n")
            continue
        top_changes = sorted(ch.get('rank_changes', []), key=lambda x: -abs(x['change']))[:8]
        sections.append(f"""
[{name}선] ({mode}) {ch['period_label']}
{warning if warning else ''}
- 신규 진입: {len(ch.get('new_entries', []))}개
- 이탈: {len(ch.get('dropped', []))}개
- 큰 변동: {len(ch.get('rank_changes', []))}개
상위 변동: {json.dumps(top_changes, ensure_ascii=False)}
""")

    prompt = f"""한국 App Store 게임 {chart_used} 차트의 다중 시간축 변화를 사업PM 관점에서 분석.

[이번 측정 Top 30]
{json.dumps(current[:30], ensure_ascii=False, indent=2)}

{"".join(sections)}

**핵심 관점**: 단기 vs 중장기 추세의 일치/불일치로 진짜 시그널 식별.
- 단기(1일·1주)만 변동, 중장기 안정 → 일시적 노이즈
- 단기·중장기 모두 변동 → 진짜 추세
- 중장기 변동 크고 단기 안정 → 추세 안착

**데이터 부족 시간축**: 비교 불가로 표시된 시간축은 분석에서 제외하고 가용 데이터 위주로 판단.

한국어 5~8문단 (각 2~4줄):
1. 활성 시간축 요약 (비교 가능/불가 여부 포함)
2. 단기 시그널 (1일선)
3. 중기 시그널 (1주·1달선, 가용 시)
4. 장기 시그널 (분기·1년선, 가용 시)
5. 단기 vs 중장기 비교
6. 사업PM 액션 포인트

군더더기 없이.""" + INDUSTRY_VOICE

    result = call_claude_with_retry(prompt, max_tokens=2500)
    if result is None:
        return ("⚠️ AI 종합 인사이트 생성 실패 — Claude API 일시적 과부하(529). "
                "시간축별 변화 데이터는 첨부 엑셀에서 정상 확인 가능합니다.")
    return result


# ============================================================
# 7b. 인사이트 캐시 (같은 기간엔 같은 인사이트 재사용)
# ============================================================

INSIGHT_CACHE_FILE = DATA_DIR / 'insight_cache.json'


def load_insight_cache():
    if INSIGHT_CACHE_FILE.exists():
        try:
            return json.loads(INSIGHT_CACHE_FILE.read_text(encoding='utf-8'))
        except Exception:
            return {}
    return {}


def save_insight_cache(cache):
    INSIGHT_CACHE_FILE.write_text(json.dumps(cache, ensure_ascii=False, indent=2), encoding='utf-8')


def _insight_signature(ch):
    """해당 시간축 비교 데이터의 지문 — 데이터가 바뀌면 인사이트 재생성."""
    payload = json.dumps({
        'pv': 'v4-industry',  # 프롬프트 버전 — 바꾸면 캐시 무효화·전체 재생성
        'new': ch.get('new_entries', []),
        'drop': ch.get('dropped', []),
        'chg': ch.get('rank_changes', []),
        'pl': ch.get('period_label', ''),
    }, ensure_ascii=False, sort_keys=True)
    return hashlib.sha256(payload.encode('utf-8')).hexdigest()[:16]


def generate_timeframe_insight(name, ch, chart_used):
    """시간축 한 개의 인사이트(Opus 4.8·적응형 사고). 비교 불가면 고정 문구."""
    if not ch.get('comparable'):
        return f"비교 가능한 데이터가 아직 부족합니다. 기간이 더 쌓이면 {name}선 분석이 시작됩니다."
    top_changes = sorted(ch.get('rank_changes', []), key=lambda x: -abs(x['change']))[:15]
    prompt = f"""한국 App Store 게임 {chart_used} 차트의 [{name}선] 변화를 사업PM 관점에서 분석.

[비교 구간] {ch['period_label']}

[신규 진입]
{json.dumps(ch.get('new_entries', [])[:10], ensure_ascii=False, indent=2)}

[차트 이탈]
{json.dumps(ch.get('dropped', [])[:10], ensure_ascii=False, indent=2)}

[큰 변동]
{json.dumps(top_changes, ensure_ascii=False, indent=2)}

아래 **여섯 개의 소제목을 정확히 이 제목·이 순서로** 써서 한국어로 작성한다(각 1~2문장, 핵심만):

## 핵심 변동
{name}선에서 가장 주목할 변동 1~2건과 의미. 상시 체류 캐시카우·데이터 누락 등 과대해석 위험은 여기서 한 줄로만 짚는다.

## 이벤트·업데이트 효과
순위를 움직인 라이브 이벤트·업데이트·콜라보를 게임명과 함께 추정(근거 약하면 솔직히).

## 지속 신호
일시적 스파이크가 아니라 꾸준히 주목할 흐름. 차트 체류일수(days_in_chart)·연속 상승 등을 근거로 {name}선 기준 지속성을 본다.

## 장르 신호
신규 진입·급등 게임의 장르(genre) 분포로 어느 장르가 뜨거나 식는지. 뚜렷하지 않으면 솔직히.

## 벤치마크 대상
오늘 가장 연구할 가치 있는 게임/무브 딱 하나와 그 이유.

## 사업PM 인사이트
위를 종합한 핵심 통찰 한 줄과 함의.

[형식 규칙] 매 시간축 동일하게 위 여섯 소제목만 사용한다. 수평선(---)·이모지·다른 소제목은 쓰지 않는다. 굵게(**)는 게임명·수치에만. 군더더기 없이.""" + INDUSTRY_VOICE
    result = call_claude_with_retry(prompt)
    if result is None:
        return "⚠️ AI 인사이트 생성 일시 실패(과부하). 변화 데이터는 섹션/첨부 엑셀에서 정상 확인 가능합니다."
    return result.strip()


def attach_insights(analyses, chart_used, cache):
    """표시할 각 시간축에 인사이트를 붙인다. 같은 기간·데이터면 캐시 재사용, 아니면 생성·저장."""
    for name, ch in analyses.items():
        key = ch.get('period_key', '')
        sig = _insight_signature(ch)
        cached = cache.get(name)
        valid_cache = (cached and cached.get('key') == key and cached.get('sig') == sig
                       and '⚠️ AI 인사이트' not in cached.get('text', ''))  # 실패 문구는 캐시 무시·재시도
        if valid_cache:
            ch['insight'] = cached['text']
            ch['insight_cached'] = True
        else:
            text = ''  # [비용절감] 레거시 /index 일일 AI 인사이트 비활성화 — 주력은 markets. AI 호출 생략(~$12/월↓). 되살리려면 generate_timeframe_insight(name, ch, chart_used) 복원.
            ch['insight'] = text
            ch['insight_cached'] = False
            cache[name] = {'key': key, 'sig': sig, 'text': text}
    return analyses


# ============================================================
# 8. 엑셀
# ============================================================

def _write_changes_to_sheet(ws, changes, title):
    is_ma = changes.get('mode') == 'moving_average'
    rank_label = '평균 순위' if is_ma else '순위'

    row = 1
    ws.cell(row=row, column=1, value=title).font = Font(bold=True, size=12)
    row += 1
    if changes.get('warning'):
        ws.cell(row=row, column=1, value=changes['warning']).font = Font(italic=True, color='C2410C')
        row += 1
    row += 1

    if not changes.get('comparable'):
        ws.cell(row=row, column=1, value='비교 불가 (데이터 부족)').font = Font(italic=True, color='888888')
        for col_letter, width in [('A', 35), ('B', 25), ('C', 40)]:
            ws.column_dimensions[col_letter].width = width
        return

    ws.cell(row=row, column=1, value='■ 신규 진입').font = Font(bold=True, size=11)
    row += 1
    items = changes.get('new_entries', [])[:30]
    if not items:
        ws.cell(row=row, column=1, value='  (해당 없음)').font = Font(italic=True, color='888888')
        row += 1
    else:
        for item in items:
            ws.cell(row=row, column=1, value=item['title'])
            ws.cell(row=row, column=2, value=item['developer'])
            if is_ma:
                ws.cell(row=row, column=3, value=f"{rank_label} {item['avg_rank']} ({item['days_in_chart']}/{item['total_days']}일 등장)")
            else:
                ws.cell(row=row, column=3, value=f"{item['rank']}위 진입")
            row += 1
    row += 2

    for label, sort_key, filter_fn in [
        ('■ 큰 폭 상승', lambda x: -x['change'], lambda x: x['change'] > 0),
        ('■ 큰 폭 하락', lambda x: x['change'], lambda x: x['change'] < 0),
    ]:
        ws.cell(row=row, column=1, value=label).font = Font(bold=True, size=11)
        row += 1
        items = [c for c in sorted(changes.get('rank_changes', []), key=sort_key)[:30] if filter_fn(c)]
        if not items:
            ws.cell(row=row, column=1, value='  (해당 없음)').font = Font(italic=True, color='888888')
            row += 1
        else:
            for item in items:
                ws.cell(row=row, column=1, value=item['title'])
                ws.cell(row=row, column=2, value=item['developer'])
                arrow = '▲' if item['change'] > 0 else '▼'
                ws.cell(row=row, column=3, value=f"{rank_label} {item['prev_rank']} → {item['curr_rank']} ({arrow}{abs(item['change'])})")
                row += 1
        row += 2

    ws.cell(row=row, column=1, value='■ 차트 이탈').font = Font(bold=True, size=11)
    row += 1
    items = changes.get('dropped', [])[:30]
    if not items:
        ws.cell(row=row, column=1, value='  (해당 없음)').font = Font(italic=True, color='888888')
        row += 1
    else:
        for item in items:
            ws.cell(row=row, column=1, value=item['title'])
            ws.cell(row=row, column=2, value=item['developer'])
            if is_ma:
                ws.cell(row=row, column=3, value=f"이전 {rank_label} {item['avg_rank']}에서 이탈")
            else:
                ws.cell(row=row, column=3, value=f"이전 {item['rank']}위에서 이탈")
            row += 1

    for col_letter, width in [('A', 35), ('B', 25), ('C', 45)]:
        ws.column_dimensions[col_letter].width = width


def create_comprehensive_excel_report(current, analyses, summary, chart_used):
    today = datetime.now().strftime('%Y%m%d')
    filename = f'mobile_chart_{today}.xlsx'
    wb = Workbook()

    ws = wb.active
    ws.title = '요약'
    ws['A1'] = f'한국 App Store 게임 차트 종합 보고서 ({datetime.now().strftime("%Y-%m-%d")})'
    ws['A1'].font = Font(bold=True, size=14)
    ws.merge_cells('A1:D1')
    ws['A2'] = f'사용 차트: {chart_used}'
    ws['A2'].font = Font(italic=True, size=11)

    ws['A4'] = '활성 분석 시간축'
    ws['A4'].font = Font(bold=True, size=12)
    lines = []
    for name, ch in analyses.items():
        warning_suffix = f"  {ch['warning']}" if ch.get('warning') else ""
        comparable_mark = "" if ch.get('comparable') else "  [비교 불가]"
        lines.append(f"  · {name}선: {ch['period_label']}{comparable_mark}{warning_suffix}")
    ws['A5'] = "\n".join(lines)
    ws['A5'].alignment = Alignment(wrap_text=True, vertical='top')
    ws.row_dimensions[5].height = max(30, len(analyses) * 30)

    ws['A7'] = 'Claude 종합 인사이트'
    ws['A7'].font = Font(bold=True, size=12)
    ws['A8'] = summary
    ws['A8'].alignment = Alignment(wrap_text=True, vertical='top')
    ws.column_dimensions['A'].width = 100
    ws.row_dimensions[8].height = 600

    ws2 = wb.create_sheet('이번 차트')
    df = pd.DataFrame(current)
    if not df.empty:
        for r in dataframe_to_rows(df, index=False, header=True):
            ws2.append(r)
        for cell in ws2[1]:
            cell.font = Font(bold=True)
            cell.fill = PatternFill(start_color='DDDDDD', end_color='DDDDDD', fill_type='solid')
        for col_letter, width in [('A', 8), ('B', 30), ('C', 40), ('D', 25), ('E', 20), ('F', 15), ('G', 20)]:
            ws2.column_dimensions[col_letter].width = width

    for name, ch in analyses.items():
        ws_t = wb.create_sheet(f'{name}선')
        title = f'■ {name}선 ({ch["period_label"]})'
        _write_changes_to_sheet(ws_t, ch, title)

    wb.save(filename)
    return filename


# ============================================================
# 9. 메일 본문
# ============================================================

def _build_section_html(name, ch):
    """시간축 한 개의 HTML 섹션."""
    is_ma = ch.get('mode') == 'moving_average'
    rank_label = '평균 순위' if is_ma else '순위'

    header = f"<h3 style='margin-top:32px; border-bottom:1px solid #ddd; padding-bottom:6px;'>📊 {name}선</h3>"
    header += f"<p style='color:#666; font-size:13px; margin:8px 0;'>{ch['period_label']}</p>"
    if ch.get('warning'):
        header += f"<p style='color:#C2410C; background:#FEF3F2; padding:8px 12px; border-radius:4px; font-size:13px;'>{ch['warning']}</p>"

    if not ch.get('comparable'):
        msg = ch.get('insight') or "비교 불가 — 데이터 부족."
        return header + f"<p style='color:#888;'>{msg}</p>"

    new_html = "".join([
        f"<li>{e['title']} <span style='color:#888'>({e['developer']}, "
        + (f"{rank_label} {e['avg_rank']}, {e['days_in_chart']}/{e['total_days']}일 등장" if is_ma else f"{e['rank']}위 진입")
        + ")</span></li>"
        for e in ch.get('new_entries', [])[:10]
    ]) or "<li style='color:#888'>없음</li>"

    dropped_html = "".join([
        f"<li>{e['title']} <span style='color:#888'>("
        + (f"이전 {rank_label} {e['avg_rank']}" if is_ma else f"이전 {e['rank']}위")
        + " → 이탈)</span></li>"
        for e in ch.get('dropped', [])[:10]
    ]) or "<li style='color:#888'>없음</li>"

    rank_sorted = sorted(ch.get('rank_changes', []), key=lambda x: -abs(x['change']))[:10]
    changes_html = "".join([
        f"<li>{c['title']} <span style='color:{'#16a34a' if c['change']>0 else '#dc2626'}'>"
        f"{rank_label} {c['prev_rank']} → {c['curr_rank']} "
        f"({'▲' if c['change']>0 else '▼'}{abs(c['change'])})</span></li>"
        for c in rank_sorted
    ]) or "<li style='color:#888'>없음</li>"

    body = f"""
    <p style='margin-top:12px; margin-bottom:4px;'><strong>📈 신규 진입</strong></p>
    <ul style='margin-top:4px;'>{new_html}</ul>
    <p style='margin-top:12px; margin-bottom:4px;'><strong>📉 차트 이탈</strong></p>
    <ul style='margin-top:4px;'>{dropped_html}</ul>
    <p style='margin-top:12px; margin-bottom:4px;'><strong>📊 큰 폭 변동</strong></p>
    <ul style='margin-top:4px;'>{changes_html}</ul>
    """
    insight = ch.get('insight')
    insight_html = ""
    if insight:
        insight_html = (
            "<div style='margin-top:12px; background:#f8f8f8; border-left:3px solid #2563eb; "
            "padding:10px 14px; border-radius:4px;'>"
            "<strong>💡 인사이트</strong>"
            "<pre style='white-space:pre-wrap; font-family:inherit; line-height:1.7; margin:6px 0 0;'>"
            f"{insight}</pre></div>"
        )
    return header + body + insight_html


def _dashboard_button_html():
    """모든 메일 상단에 들어가는 대시보드 링크 버튼."""
    return (
        f"<p style='margin:16px 0 4px;'>"
        f"<a href='{DASHBOARD_URL}' "
        f"style='display:inline-block; background:#2563eb; color:#ffffff; text-decoration:none; "
        f"padding:11px 20px; border-radius:8px; font-weight:600; font-size:15px;'>"
        f"📊 인터랙티브 차트 대시보드 열기 →</a></p>"
        f"<p style='color:#888; font-size:12px; margin:0 0 8px;'>"
        f"일·주·월·분기·년 시간축 · 100여 개 게임 추이 · 보고 싶은 게임만 선택해 강조</p>"
    )


def build_daily_only_email_html(today, chart_used, current, daily_changes, summary):
    """평일 일간 라이트 본문."""
    section = _build_section_html('1일', daily_changes)
    return f"""
    <div style="font-family: 'Malgun Gothic', sans-serif; line-height: 1.6; max-width: 700px;">
      <h2>📱 한국 App Store 게임 차트 일간 변동</h2>
      <p><strong>수집일:</strong> {today} | <strong>차트:</strong> {chart_used} ({len(current)}개)</p>
      {_dashboard_button_html()}
      <hr/>
      {section}
      <h3 style="margin-top:32px;">💡 인사이트</h3>
      <pre style="white-space: pre-wrap; font-family: 'Malgun Gothic', sans-serif; line-height: 1.7; background: #f8f8f8; padding: 16px; border-radius: 4px;">{summary}</pre>
      <hr style="margin-top: 30px;"/>
      <p style="color: #888; font-size: 12px;">월요일에는 1주선이, 매월 1일에는 1달선이, 분기 시작일에는 분기선이, 1/1에는 1년선이 추가됩니다.</p>
    </div>
    """


def build_comprehensive_email_html(today, chart_used, current, analyses, summary):
    """종합 모드 본문 (시간축 여러 개)."""
    sections_html = "".join([_build_section_html(name, ch) for name, ch in analyses.items()])

    tf_summary = "<ul>"
    for name, ch in analyses.items():
        mark = "✅" if ch.get('comparable') else "⚠️"
        tf_summary += f"<li>{mark} <strong>{name}선</strong> — {ch['period_label']}</li>"
    tf_summary += "</ul>"

    return f"""
    <div style="font-family: 'Malgun Gothic', sans-serif; line-height: 1.6; max-width: 800px;">
      <h2>📱 한국 App Store 게임 차트 종합 보고서</h2>
      <p><strong>수집일:</strong> {today} | <strong>차트:</strong> {chart_used} ({len(current)}개)</p>
      {_dashboard_button_html()}
      <hr/>
      <h3>📋 이번 보고 활성 시간축</h3>
      {tf_summary}
      {sections_html}
      <p style="color: #666; margin-top: 24px;">각 시간축 인사이트는 위 섹션에, 상세 데이터는 첨부 엑셀에 있습니다. 같은 기간의 인사이트는 데이터가 바뀔 때까지 동일하게 유지됩니다.</p>
    </div>
    """


# ============================================================
# 10. 메일 발송
# ============================================================

def send_email_via_gmail(subject, html_body, attachment_path=None):
    msg = MIMEMultipart()
    msg['From'] = GMAIL_USER
    msg['To'] = RECIPIENT_EMAIL
    msg['Subject'] = subject
    msg.attach(MIMEText(html_body, 'html', 'utf-8'))
    if attachment_path:
        with open(attachment_path, 'rb') as f:
            part = MIMEBase('application', 'octet-stream')
            part.set_payload(f.read())
        encoders.encode_base64(part)
        part.add_header('Content-Disposition', f'attachment; filename={os.path.basename(attachment_path)}')
        msg.attach(part)
    app_password = GMAIL_APP_PASSWORD.replace(' ', '')
    with smtplib.SMTP_SSL('smtp.gmail.com', 465) as server:
        server.login(GMAIL_USER, app_password)
        server.send_message(msg)
    print(f"[OK] 메일 발송: {RECIPIENT_EMAIL}")


# ============================================================
# 11. 메인 (매일 종합 + 시간축별 캐시 인사이트)
# ============================================================

def main():
    print(f"\n=== 한국 App Store 게임 차트 수집 ({datetime.now()}) ===\n")

    missing = [k for k, v in {
        'ANTHROPIC_API_KEY': ANTHROPIC_API_KEY, 'GMAIL_USER': GMAIL_USER,
        'GMAIL_APP_PASSWORD': GMAIL_APP_PASSWORD, 'RECIPIENT_EMAIL': RECIPIENT_EMAIL,
    }.items() if not v]
    if missing:
        raise RuntimeError(f"환경변수 누락: {missing}")

    print(f"[INFO] AI 분석 모델: {CLAUDE_MODEL} (적응형 사고·effort={THINKING_EFFORT})")
    today_dt = datetime.now()

    print("[1/4] App Store 차트 수집...")
    current, chart_used = fetch_apple_chart_kr_games(100)
    if not current:
        raise RuntimeError("모든 차트 수집 실패")
    print(f"      → {len(current)}개 수집, 사용 차트: {chart_used}")
    attach_genres(current)

    print("[2/4] 전체 시간축 분석...")
    all_analyses = run_active_analyses(today_dt, current, get_active_timeframes(today_dt))
    # 표시 대상: 1일은 항상, 상위 시간축은 비교 가능할 때만(데이터 쌓이면 자동 등장)
    display_names = ['1일'] + [n for n in ['1주', '1달', '분기', '1년']
                               if all_analyses.get(n, {}).get('comparable')]
    analyses = {n: all_analyses[n] for n in display_names}
    for name, ch in analyses.items():
        mark = "✅" if ch.get('comparable') else "⚠️"
        print(f"      · {mark} {name}선: {ch['period_label']}")

    # 인사이트용: 현재 차트 장르를 비교 항목에 보강(장르 신호 섹션용)
    genre_map = {it.get('title'): it.get('genre', '미상') for it in current if it.get('title')}
    for _ch in analyses.values():
        for _k in ('new_entries', 'dropped', 'rank_changes'):
            for _it in (_ch.get(_k) or []):
                _it.setdefault('genre', genre_map.get(_it.get('title'), '미상'))

    today = today_dt.strftime('%Y-%m-%d')

    print("[3/4] 시간축별 인사이트 (같은 기간이면 캐시 재사용)...")
    cache = load_insight_cache()
    analyses = attach_insights(analyses, chart_used, cache)
    save_insight_cache(cache)
    for name, ch in analyses.items():
        print(f"      · {name}선 인사이트: {'캐시 재사용' if ch.get('insight_cached') else '신규 생성'}")
    combined = "\n\n".join(f"[{name}선] {ch['period_label']}\n{ch['insight']}"
                           for name, ch in analyses.items())
    print("─" * 60)
    print(combined)
    print("─" * 60)

    # 대시보드용 AI 브리핑 저장(날짜별 누적 히스토리, 집계 스크립트가 data.json에 포함)
    try:
        brief_path = DATA_DIR / 'ai_brief.json'
        existing = {}
        if brief_path.exists():
            try:
                existing = json.loads(brief_path.read_text(encoding='utf-8'))
            except Exception:
                existing = {}
        entries = existing.get('entries', []) if isinstance(existing, dict) else []
        today_date = today_dt.strftime('%Y-%m-%d')
        entry = {
            'date': today_date,
            'generated': today_dt.strftime('%Y-%m-%d %H:%M'),
            'chart': chart_used,
            'items': [{'name': name, 'period': ch.get('period_label', ''), 'text': ch.get('insight', ''),
                       'counts': {'new': len(ch.get('new_entries', []) or []),
                                  'drop': len(ch.get('dropped', []) or []),
                                  'up': len([c for c in (ch.get('rank_changes', []) or []) if (c.get('change', 0) or 0) > 0]),
                                  'down': len([c for c in (ch.get('rank_changes', []) or []) if (c.get('change', 0) or 0) < 0])}}
                      for name, ch in analyses.items()],
        }
        entries = [e for e in entries if e.get('date') != today_date]
        entries.insert(0, entry)
        entries = entries[:30]
        brief_path.write_text(json.dumps({'updated': entry['generated'], 'entries': entries},
                                         ensure_ascii=False, indent=2), encoding='utf-8')
        print(f'[OK] AI 브리핑 저장: data/ai_brief.json (entries={len(entries)})')
    except Exception as e:
        print(f'[WARN] AI 브리핑 저장 실패: {e}')

    print("[4/4] 엑셀·메일 발송 (매일 종합)...")
    excel_path = create_comprehensive_excel_report(current, analyses, combined, chart_used)
    subject = f'[모바일 게임 차트] 종합 보고 {today} ({chart_used})'
    html_body = build_comprehensive_email_html(today, chart_used, current, analyses, combined)
    send_email_via_gmail(subject, html_body, attachment_path=excel_path)

    save_current_data(current)

    print("[보조] Top Free(인기) 차트 수집·저장...")
    free = fetch_top_free_kr_games(100)
    if free:
        attach_genres(free)
        save_free_data(free)

    print("[다국가] 전체 국가 iOS 차트 수집·저장 + 주요시장/지역/글로벌 브리핑...")
    try:
        collected = collect_all_countries(COUNTRIES + COUNTRIES_EXTRA)  # [임시·1회] 전 32개국 수집→전 지역 브리핑 생성(이후 코어10로 복귀, 지역탭은 persist 로직으로 유지)
        majors = build_major_briefs(collected)      # 주요 시장 개별 먼저
        regions = build_regional_briefs(collected)  # 중소규모 지역 묶음(전 지역 생성)
        write_briefs_index(majors, regions)         # 탭 인덱스(주요→지역)
        build_global_brief(collected)               # 주요+지역을 토대로 글로벌 종합
        apply_aliases_to_briefs()                   # 브리핑 내 현지어 게임명→한글 치환(무비용)
    except Exception as e:
        print(f"[WARN] 다국가 수집/브리핑 실패: {e}")

    print("\n=== 완료 ===\n")


if __name__ == '__main__':
    main()
